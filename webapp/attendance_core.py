"""
考勤核心处理模块 v2 — 支持动态天数
"""
import re
import io
from typing import Tuple, List, Dict, Optional, BinaryIO
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

R_NAME, R_ACCOUNT, R_CLOCK_START = 0, 1, 5
M_NAME, M_ACCOUNT = 0, 1
M_DAILY_START = 46
SUMMARY_COLS = 6  # 出勤, 休息, 事假, 迟到/缺卡, 缺卡, 备注


def _time_to_min(t: Optional[str]) -> Optional[int]:
    if not t: return None
    try:
        h, m = t.strip().split(':')
        return int(h) * 60 + int(m)
    except (ValueError, IndexError):
        return None


def parse_clock(text: str) -> Tuple[Optional[str], Optional[str], int]:
    if not text or str(text).strip() in ('', '--', 'None'):
        return None, None, 0
    times = re.findall(r'(\d{2}:\d{2})', str(text))
    if not times:
        return None, None, 0
    if len(times) == 1:
        t = times[0]
        return (t, None, 1) if (_time_to_min(t) or 0) < 12 * 60 else (None, t, 1)
    return times[0], times[-1], len(times)


def determine_status(clock_cell: str, monthly_cell: str) -> Tuple[str, str, bool]:
    in_time, out_time, punches = parse_clock(clock_cell)
    m_text = str(monthly_cell).strip() if monthly_cell else ''

    if '病假' in m_text:
        return 'sick_leave', '△', False
    for kw in ['请假', '事假', '年假', '调休', '婚假', '产假', '陪产假', '丧假']:
        if kw in m_text:
            return 'leave', '×', False

    if '未排班' in m_text or '休息' in m_text:
        # 休息日有打卡 → 加班，不计迟到早退
        if punches > 0:
            return 'overtime', '√', False
        return 'rest', '休', True

    if punches == 0 and (not m_text or m_text in ('--', 'None')):
        return 'empty', '', False

    if punches > 0:
        inm = _time_to_min(in_time)
        outm = _time_to_min(out_time)
        # 班次判定: 上班<12:00→早班(9-18), 上班>=12:00→晚班(13-22) 严格卡死
        is_evening = inm is not None and inm >= 12 * 60
        is_late = False
        is_early = False
        if inm is not None:
            if is_evening:
                is_late = inm >= 13 * 60
                if outm is not None:
                    is_early = outm < 22 * 60
            else:
                is_late = inm >= 9 * 60
                if outm is not None:
                    is_early = outm < 18 * 60
        is_miss = punches == 1
        is_device = '设备异常' in m_text

        if is_late: return 'late', '≦', False
        if is_miss: return 'miss_punch', '缺卡', False
        if is_early: return 'early', '早退', False
        if is_device: return 'device_error', '设备异常', False
        if punches >= 2: return 'normal', '√', False
        return 'unknown', '?', False

    for kw, st, lb in [('缺卡', 'miss_punch', '缺卡'), ('迟到', 'late', '≦'),
                        ('设备异常', 'device_error', '设备异常'),
                        ('正常', 'normal', '√'), ('旷工', 'absent', 'Ο')]:
        if kw in m_text:
            return st, lb, False
    return 'empty', '', False


def detect_employment(daily_clock: List[str], daily_m: List[str]) -> Tuple[int, int]:
    n = len(daily_clock)
    first, has_any = 0, False
    last = n - 1
    for d in range(n):
        c = daily_clock[d] if d < len(daily_clock) else ''
        m = daily_m[d] if d < len(daily_m) else ''
        c_ok = c and c not in ('--', 'None', '')
        m_ok = m and m not in ('--', 'None', '') and '未排班' not in m and '休息' not in m
        if c_ok or m_ok:
            if not has_any:
                first = d
                has_any = True
            last = d
    return (first, last) if has_any else (0, -1)


def process(record_stream: BinaryIO, monthly_stream: BinaryIO):
    wb_r = load_workbook(record_stream); ws_r = wb_r.active
    wb_m = load_workbook(monthly_stream); ws_m = wb_m.active

    # 检测天数：打卡记录从 R_CLOCK_START 列到最后一列
    num_days = ws_r.max_column - R_CLOCK_START

    # ── 读取月报 ──
    monthly_data = {}
    for row_idx in range(5, ws_m.max_row + 1):
        def m_val(c): return ws_m.cell(row=row_idx, column=c + 1).value
        name = str(m_val(M_NAME) or '').strip()
        account = str(m_val(M_ACCOUNT) or '').strip()
        if not account or not name:
            continue
        leave_raw = m_val(29)
        try:
            leave_days = float(leave_raw) if leave_raw and str(leave_raw).strip() not in ('', '--', 'None') else 0
        except (ValueError, TypeError):
            leave_days = 0
        daily_m = [str(ws_m.cell(row=row_idx, column=M_DAILY_START + d + 1).value or '')
                   for d in range(num_days)]
        monthly_data[account.lower()] = {'name': name, 'account': account,
                                         'leave_days': leave_days, 'daily_m': daily_m}

    # ── 读取打卡记录 ──
    persons_raw = []
    for row_idx in range(5, ws_r.max_row + 1):
        def r_val(c): return ws_r.cell(row=row_idx, column=c + 1).value
        name = str(r_val(R_NAME) or '').strip()
        account = str(r_val(R_ACCOUNT) or '').strip()
        if not name:
            continue
        clock = [str(ws_r.cell(row=row_idx, column=R_CLOCK_START + d + 1).value or '')
                 for d in range(num_days)]
        mi = monthly_data.get(account.lower(), {})
        persons_raw.append({'name': name, 'account': account, 'clock': clock,
                            'daily_m': mi.get('daily_m', [''] * num_days),
                            'leave_days': mi.get('leave_days', 0)})

    # 补充月报有但打卡记录无的人
    existing = {p['account'].lower() for p in persons_raw}
    for acct, mi in monthly_data.items():
        if acct not in existing:
            persons_raw.append({'name': mi['name'], 'account': mi['account'],
                                'clock': [''] * num_days, 'daily_m': mi['daily_m'],
                                'leave_days': mi['leave_days']})

    # 计算当月休息天数
    ref_work = 0; month_rest = 0
    for p in persons_raw:
        w = sum(1 for d in p['daily_m'] if '正常' in d or '迟到' in d or '缺卡' in d)
        r = sum(1 for d in p['daily_m'] if '未排班' in d)
        if w > ref_work:
            ref_work = w; month_rest = r
    if month_rest == 0:
        month_rest = num_days - ref_work if ref_work else 5

    # ── 逐人生成 ──
    results = []; total_late = 0; total_miss = 0; total_leave = 0

    for p in persons_raw:
        emp_first, emp_last = detect_employment(p['clock'], p['daily_m'])

        daily_statuses = []
        work = late = miss = early = rest = 0

        for d in range(num_days):
            if d < emp_first or d > emp_last:
                daily_statuses.append(''); continue

            c = p['clock'][d] if d < len(p['clock']) else ''
            m = p['daily_m'][d] if d < len(p['daily_m']) else ''
            stype, display, is_rest = determine_status(c, m)

            if stype == 'rest': rest += 1
            elif stype == 'overtime': work += 1
            elif stype in ('normal', 'device_error', 'business'): work += 1
            elif stype == 'late': work += 1; late += 1
            elif stype == 'miss_punch': work += 1; miss += 1
            elif stype == 'early': work += 1; early += 1

            daily_statuses.append(display)

        # 排除无任何出勤数据的人员
        has_any_work = work > 0 or late > 0 or miss > 0 or p['leave_days'] > 0
        if not has_any_work:
            continue

        late_col = late if late > 0 else ''
        miss_col = miss if miss > 0 else ''

        remark_parts = []
        if late > 0: remark_parts.append(f'迟到{late}次')
        if miss > 0: remark_parts.append(f'缺卡{miss}次')
        if late + miss + early >= 3: remark_parts.append('迟到/缺卡扣20元')

        total_late += late; total_miss += miss; total_leave += p['leave_days']

        results.append({
            'name': p['name'], 'account': p['account'],
            'daily_statuses': daily_statuses,
            'actual_attendance': work, 'rest_days': month_rest,
            'leave_days': p['leave_days'] if p['leave_days'] > 0 else '',
            'late': late_col, 'miss_punch': miss_col,
            'remarks': '; '.join(remark_parts),
        })

    # 星期标头
    weekdays = []
    for d in range(num_days):
        val = ws_m.cell(row=4, column=M_DAILY_START + d + 1).value
        wd = str(val).strip() if val else ''
        parts = wd.split('\n')
        weekdays.append(parts[-1] if len(parts) > 1 else wd)

    return {
        'persons': results, 'weekdays': weekdays,
        'month_rest_days': month_rest, 'num_days': num_days,
        'summary': {
            'total_late': total_late, 'total_miss': total_miss,
            'total_leave': total_leave, 'person_count': len(results),
        }
    }


def generate_excel(results: List[Dict], weekdays: List[str]) -> io.BytesIO:
    """基于模板文件生成考勤 Excel"""
    from pathlib import Path

    template_path = Path(__file__).resolve().parent / 'template.xlsx'
    if not template_path.exists():
        # Fallback: generate from scratch if no template
        return _generate_excel_scratch(results, weekdays)

    wb = load_workbook(template_path)
    ws = wb.active
    num_days = len(weekdays)

    # 从 Row4 检测模板结构：找到第一个汇总标题列
    tpl_day_end = 3  # 默认
    tpl_summary_start = None
    summary_labels = {'实际出勤天数(天)', '休息(天)', '事假(天)', '迟到/缺卡次数(次)', '缺卡次数(次)', '备注'}
    for col in range(4, ws.max_column + 1):
        v = str(ws.cell(row=4, column=col).value or '').replace('\n', '')
        if v in summary_labels:
            tpl_summary_start = col
            tpl_day_end = col - 1
            break

    if tpl_summary_start is None:
        tpl_day_end = ws.max_column - 6  # fallback
        tpl_summary_start = tpl_day_end + 1

    template_days = tpl_day_end - 3

    day_end_col = 3 + num_days
    summary_start_col = day_end_col + 1

    # 如果实际天数超过模板天数，在汇总列前插入日列
    if num_days > template_days:
        extra = num_days - template_days
        for _ in range(extra):
            ws.insert_cols(tpl_summary_start)
        # summary_start 不变（已在代码中基于 num_days 计算）

    # Row 1, Row 2 保持模板原文不修改
    # 更新 Row 4 日标头
    for day in range(1, num_days + 1):
        wd = weekdays[day - 1] if day - 1 < len(weekdays) else ''
        cell = ws.cell(row=4, column=3 + day)
        cell.value = f'{day}\n{wd}'

    # 颜色填充
    from openpyxl.styles import PatternFill
    fills = {
        '≦': PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid'),
        '缺卡': PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid'),
        '×': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        '△': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
    }

    # 复制 Row4 样式作为数据行样式模板
    ref_cell = ws.cell(row=4, column=4)

    # 清空旧数据行并填入新数据
    # 先删除模板中的示例数据行（row 5 及之后）
    last_data_row = ws.max_row
    if last_data_row > 4:
        ws.delete_rows(5, last_data_row - 4)

    for idx, r in enumerate(results):
        row = 5 + idx
        ws.cell(row=row, column=1).value = idx + 1
        ws.cell(row=row, column=2).value = r['name']
        ws.cell(row=row, column=3).value = r['account']
        for d, status in enumerate(r['daily_statuses']):
            cell = ws.cell(row=row, column=4 + d)
            cell.value = status if status else ''
            if status in fills:
                cell.fill = fills[status]
        ws.cell(row=row, column=summary_start_col).value = r['actual_attendance']
        ws.cell(row=row, column=summary_start_col + 1).value = r['rest_days']
        ws.cell(row=row, column=summary_start_col + 2).value = r['leave_days']
        ws.cell(row=row, column=summary_start_col + 3).value = r['late']
        ws.cell(row=row, column=summary_start_col + 4).value = r['miss_punch']
        ws.cell(row=row, column=summary_start_col + 5).value = r['remarks']

    # 数据行统一字体：楷体 10号 居中（第5行起，前4行保持模板表头原样）
    kaiti_font = Font(name='楷体', size=10)
    ka_align = Alignment(horizontal='center', vertical='center')
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = kaiti_font
            cell.alignment = ka_align

    ws.freeze_panes = 'D5'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _generate_excel_scratch(results: List[Dict], weekdays: List[str]) -> io.BytesIO:
    """无模板时的备用生成"""
    num_days = len(weekdays)
    total_cols = 3 + num_days + SUMMARY_COLS
    day_end_col = 3 + num_days
    summary_start_col = day_end_col + 1

    wb = Workbook(); ws = wb.active; ws.title = '上下班打卡_月报'
    font_name = '微软雅黑'
    hfont = Font(name=font_name, size=10, bold=True)
    cfont = Font(name=font_name, size=10)
    calign = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    hfill = PatternFill(start_color='FAFBFC', end_color='FAFBFC', fill_type='solid')
    fills = {
        '≦': PatternFill(start_color='FFD7D7', end_color='FFD7D7', fill_type='solid'),
        '缺卡': PatternFill(start_color='FFE0B2', end_color='FFE0B2', fill_type='solid'),
        '×': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
        '△': PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid'),
    }

    def sc(row, col, val, font=cfont, fill=None):
        c = ws.cell(row=row, column=col, value=val)
        c.font = font; c.alignment = calign; c.border = border
        if fill: c.fill = fill

    # Row 1
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    sc(1, 1, '成都善多多健康科技有限公司_月报', hfont)
    # Row 2
    tl = sum(1 for r in results for s in r['daily_statuses'] if s == '≦')
    tm = sum(1 for r in results for s in r['daily_statuses'] if s == '缺卡')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    sc(2, 1, f'统计时间   迟到:{tl}次    缺卡:{tm}次', cfont)
    # Row 3
    sc(3, 1, '序号', hfont); sc(3, 2, '姓名', hfont); sc(3, 3, '账号', hfont)
    sc(3, 4, '出勤明细', hfont)
    ws.merge_cells(start_row=3, start_column=4, end_row=3, end_column=day_end_col)
    sc(3, summary_start_col, '出勤概况', hfont)
    ws.merge_cells(start_row=3, start_column=summary_start_col, end_row=3, end_column=total_cols)
    # Row 4
    for day in range(1, num_days + 1):
        wd = weekdays[day - 1] if day - 1 < len(weekdays) else ''
        sc(4, 3 + day, f'{day}\n{wd}', hfont, hfill)
    for i, h in enumerate(['实际出勤天数(天)', '休息(天)', '事假(天)', '迟到/缺卡次数(次)', '缺卡次数(次)', '备注']):
        sc(4, summary_start_col + i, h, hfont, hfill)
    # Data
    for idx, r in enumerate(results):
        row = 5 + idx
        sc(row, 1, idx + 1); sc(row, 2, r['name']); sc(row, 3, r['account'])
        for d, status in enumerate(r['daily_statuses']):
            sc(row, 4 + d, status if status else '', fill=fills.get(status))
        sc(row, summary_start_col, r['actual_attendance'])
        sc(row, summary_start_col + 1, r['rest_days'])
        sc(row, summary_start_col + 2, r['leave_days'])
        sc(row, summary_start_col + 3, r['late'])
        sc(row, summary_start_col + 4, r['miss_punch'])
        sc(row, summary_start_col + 5, r['remarks'])

    ws.column_dimensions['A'].width = 3.75
    ws.column_dimensions['B'].width = 6.875
    ws.column_dimensions['C'].width = 10.0
    for d in range(num_days):
        ws.column_dimensions[get_column_letter(4 + d)].width = 4.625
    for i, w in enumerate([5.625]*5 + [24.875]):
        ws.column_dimensions[get_column_letter(summary_start_col + i)].width = w
    ws.row_dimensions[1].height = 27
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 24
    ws.row_dimensions[4].height = 24
    for r in range(5, 5 + len(results)):
        ws.row_dimensions[r].height = 24
    ws.freeze_panes = 'D5'

    output = io.BytesIO(); wb.save(output); output.seek(0)
    return output
